import os
import logging
from datetime import datetime, timedelta
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from backend.models.models import Attendance, Student

logger = logging.getLogger(__name__)
REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


def get_filtered_attendance(db: Session, start_date=None, end_date=None, branch=None, student_id=None):
    query = db.query(Attendance, Student).join(Student, Attendance.student_id == Student.id)
    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)
    if branch:
        query = query.filter(Student.branch == branch)
    if student_id:
        query = query.filter(Student.student_id == student_id)
    return query.order_by(Attendance.date.desc(), Student.name).all()


def generate_pdf_report(db: Session, start_date=None, end_date=None, branch=None, student_id=None) -> str:
    records = get_filtered_attendance(db, start_date, end_date, branch, student_id)
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=18, spaceAfter=12)
        elements.append(Paragraph("AI Face Recognition Attendance System", title_style))
        elements.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        if branch:
            elements.append(Paragraph(f"Branch: {branch}", styles['Normal']))
        if start_date or end_date:
            elements.append(Paragraph(f"Period: {start_date or 'All'} to {end_date or 'All'}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        headers = ["#", "Student ID", "Name", "Roll No", "Branch", "Date", "Time In", "Status", "Confidence"]
        data = [headers]
        for i, (att, stu) in enumerate(records, 1):
            data.append([
                str(i), stu.student_id, stu.name, stu.roll_number, stu.branch,
                att.date, att.time_in.strftime("%H:%M:%S") if att.time_in else "-",
                att.status.upper(), f"{att.confidence:.1f}%" if att.confidence else "Manual"
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4ff')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,1), (-1,-1), 8),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"Total Records: {len(records)}", styles['Normal']))
        doc.build(elements)
        logger.info(f"PDF report generated: {filepath}")
        return filepath
    except ImportError:
        # Fallback: simple text PDF
        with open(filepath.replace('.pdf', '.txt'), 'w') as f:
            f.write("ATTENDANCE REPORT\n")
            f.write(f"Generated: {datetime.now()}\n\n")
            for att, stu in records:
                f.write(f"{stu.student_id} | {stu.name} | {stu.branch} | {att.date} | {att.status}\n")
        return filepath.replace('.pdf', '.txt')


def generate_excel_report(db: Session, start_date=None, end_date=None, branch=None, student_id=None) -> str:
    records = get_filtered_attendance(db, start_date, end_date, branch, student_id)
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = "AI Face Recognition Attendance System - Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Branch: {branch or 'All'}"
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ["#", "Student ID", "Name", "Roll Number", "Branch", "Date", "Time In", "Status", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for i, (att, stu) in enumerate(records, 1):
            row = i + 4
            row_data = [
                i, stu.student_id, stu.name, stu.roll_number, stu.branch,
                att.date, att.time_in.strftime("%H:%M:%S") if att.time_in else "-",
                att.status.upper(), f"{att.confidence:.1f}%" if att.confidence else "Manual"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if i % 2 == 0:
                    cell.fill = alt_fill

        col_widths = [5, 15, 25, 15, 15, 12, 12, 10, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = "Summary Statistics"
        ws2['A1'].font = Font(bold=True, size=14)
        
        total = len(records)
        present = sum(1 for att, _ in records if att.status == 'present')
        branches = {}
        for att, stu in records:
            branches[stu.branch] = branches.get(stu.branch, 0) + 1

        ws2['A3'] = "Total Records"
        ws2['B3'] = total
        ws2['A4'] = "Present"
        ws2['B4'] = present
        ws2['A6'] = "Branch"
        ws2['B6'] = "Count"
        for i, (br, cnt) in enumerate(branches.items(), 7):
            ws2[f'A{i}'] = br
            ws2[f'B{i}'] = cnt

        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")
        return filepath
    except ImportError:
        logger.error("openpyxl not installed")
        raise Exception("openpyxl required for Excel reports. Install with: pip install openpyxl")
